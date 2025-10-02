import os
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from database.manager import PriceBookManager
from diff_engine import DiffEngine

class ExportManager:
    """Manager for exporting data to various formats"""
    
    def __init__(self, database_url: str = None):
        self.price_book_manager = PriceBookManager(database_url)
        self.diff_engine = DiffEngine(database_url)
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # Ensure exports directory exists
        os.makedirs('exports', exist_ok=True)
    
    def export_price_book(self, price_book_id: int, format: str = 'excel') -> str:
        """Export a price book to specified format"""
        try:
            # Get price book data
            summary = self.price_book_manager.get_price_book_summary(price_book_id)
            products = self.price_book_manager.get_products_by_price_book(price_book_id, limit=10000)
            
            if not products:
                raise ValueError("No products found for this price book")
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            manufacturer = summary['manufacturer'].replace(' ', '_')
            edition = summary['edition'] or 'Unknown'
            filename = f"{manufacturer}_{edition}_{timestamp}"
            
            if format.lower() == 'excel':
                return self._export_to_excel(price_book_id, products, summary, filename)
            elif format.lower() == 'csv':
                return self._export_to_csv(products, summary, filename)
            elif format.lower() == 'json':
                return self._export_to_json(products, summary, filename)
            else:
                raise ValueError(f"Unsupported format: {format}")
                
        except Exception as e:
            self.logger.error(f"Error exporting price book {price_book_id}: {e}")
            raise
    
    def _export_to_excel(self, price_book_id: int, products: List[Dict], summary: Dict, filename: str) -> str:
        """Export to Excel format with formatting"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create main products sheet
        products_sheet = wb.create_sheet("Products")
        self._create_products_sheet(products_sheet, products, summary)
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Summary")
        self._create_summary_sheet(summary_sheet, summary, products)
        
        # Create metadata sheet
        metadata_sheet = wb.create_sheet("Metadata")
        self._create_metadata_sheet(metadata_sheet, summary)
        
        # Save file
        filepath = os.path.join('exports', f"{filename}.xlsx")
        wb.save(filepath)
        
        return filepath
    
    def _create_products_sheet(self, sheet, products: List[Dict], summary: Dict):
        """Create formatted products sheet"""
        # Headers
        headers = ['SKU', 'Model', 'Description', 'Base Price', 'Effective Date', 'Status', 'Family']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data
        for row, product in enumerate(products, 2):
            sheet.cell(row=row, column=1, value=product['sku'])
            sheet.cell(row=row, column=2, value=product['model'] or 'N/A')
            sheet.cell(row=row, column=3, value=product['description'] or 'N/A')
            sheet.cell(row=row, column=4, value=product['base_price'])
            sheet.cell(row=row, column=5, value=product['effective_date'] or 'N/A')
            sheet.cell(row=row, column=6, value='Active' if product['is_active'] else 'Inactive')
            sheet.cell(row=row, column=7, value=product['family'] or 'N/A')
            
            # Format price column
            if product['base_price']:
                sheet.cell(row=row, column=4).number_format = '$#,##0.00'
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add table formatting
        table = Table(displayName="ProductsTable", ref=f"A1:{column_letter}{len(products) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        sheet.add_table(table)
    
    def _create_summary_sheet(self, sheet, summary: Dict, products: List[Dict]):
        """Create summary statistics sheet"""
        # Title
        sheet.cell(row=1, column=1, value=f"{summary['manufacturer']} Price Book Summary")
        sheet.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        # Statistics
        stats = [
            ("Manufacturer", summary['manufacturer']),
            ("Edition", summary['edition'] or 'N/A'),
            ("Effective Date", summary['effective_date'] or 'N/A'),
            ("Upload Date", summary['upload_date']),
            ("Status", summary['status']),
            ("Total Products", len(products)),
            ("Active Products", len([p for p in products if p['is_active']])),
            ("Inactive Products", len([p for p in products if not p['is_active']])),
        ]
        
        # Add price statistics
        prices = [p['base_price'] for p in products if p['base_price']]
        if prices:
            stats.extend([
                ("Average Price", f"${sum(prices) / len(prices):.2f}"),
                ("Highest Price", f"${max(prices):.2f}"),
                ("Lowest Price", f"${min(prices):.2f}"),
            ])
        
        # Write statistics
        for row, (label, value) in enumerate(stats, 3):
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=1).font = Font(bold=True)
            sheet.cell(row=row, column=2, value=value)
    
    def _create_metadata_sheet(self, sheet, summary: Dict):
        """Create metadata sheet"""
        metadata = [
            ("Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Price Book ID", summary['id']),
            ("Manufacturer", summary['manufacturer']),
            ("Edition", summary['edition'] or 'N/A'),
            ("Effective Date", summary['effective_date'] or 'N/A'),
            ("Upload Date", summary['upload_date']),
            ("Status", summary['status']),
            ("File Path", summary['file_path']),
        ]
        
        for row, (label, value) in enumerate(metadata, 1):
            sheet.cell(row=row, column=1, value=label)
            sheet.cell(row=row, column=1).font = Font(bold=True)
            sheet.cell(row=row, column=2, value=value)
    
    def _export_to_csv(self, products: List[Dict], summary: Dict, filename: str) -> str:
        """Export to CSV format"""
        # Create DataFrame
        df = pd.DataFrame(products)

        # Add metadata as first row
        metadata_row = {
            'SKU': f"# {summary['manufacturer']} Price Book Export",
            'Model': f"Edition: {summary['edition'] or 'N/A'}",
            'Description': f"Effective Date: {summary['effective_date'] or 'N/A'}",
            'Base Price': f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            'Effective Date': f"Total Products: {len(products)}",
            'Status': '',
            'Family': ''
        }

        # Insert metadata row
        df = pd.concat([pd.DataFrame([metadata_row]), df], ignore_index=True)

        # Save CSV
        filepath = os.path.join('exports', f"{filename}.csv")
        df.to_csv(filepath, index=False)

        return filepath

    def _export_to_json(self, products: List[Dict], summary: Dict, filename: str) -> str:
        """Export to JSON format"""
        # Build complete export structure
        export_data = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'manufacturer': summary['manufacturer'],
                'edition': summary['edition'],
                'effective_date': summary['effective_date'],
                'upload_date': summary['upload_date'],
                'price_book_id': summary['id'],
                'status': summary['status'],
                'total_products': len(products)
            },
            'products': products
        }

        # Save JSON with pretty formatting
        filepath = os.path.join('exports', f"{filename}.json")
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)

        return filepath
    
    def export_change_log(self, old_price_book_id: int, new_price_book_id: int, format: str = 'excel') -> str:
        """Export change log between two price books"""
        try:
            # Get change log
            changes = self.diff_engine.get_change_log(old_price_book_id, new_price_book_id)
            
            if not changes:
                raise ValueError("No changes found between the selected price books")
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"change_log_{old_price_book_id}_to_{new_price_book_id}_{timestamp}"
            
            if format.lower() == 'excel':
                return self._export_change_log_to_excel(changes, filename)
            elif format.lower() == 'csv':
                return self._export_change_log_to_csv(changes, filename)
            else:
                raise ValueError(f"Unsupported format: {format}")
                
        except Exception as e:
            self.logger.error(f"Error exporting change log: {e}")
            raise
    
    def _export_change_log_to_excel(self, changes: List[Dict], filename: str) -> str:
        """Export change log to Excel"""
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Change Log"
        
        # Headers
        headers = ['Change Type', 'Product ID', 'Old Value', 'New Value', 'Change %', 'Description', 'Created At']
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data
        for row, change in enumerate(changes, 2):
            sheet.cell(row=row, column=1, value=change['change_type'])
            sheet.cell(row=row, column=2, value=change['product_id'] or 'N/A')
            sheet.cell(row=row, column=3, value=change['old_value'] or 'N/A')
            sheet.cell(row=row, column=4, value=change['new_value'] or 'N/A')
            sheet.cell(row=row, column=5, value=change['change_percentage'] or 'N/A')
            sheet.cell(row=row, column=6, value=change['description'])
            sheet.cell(row=row, column=7, value=change['created_at'])
            
            # Format percentage column
            if change['change_percentage']:
                sheet.cell(row=row, column=5).number_format = '0.00%'
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        filepath = os.path.join('exports', f"{filename}.xlsx")
        wb.save(filepath)
        
        return filepath
    
    def _export_change_log_to_csv(self, changes: List[Dict], filename: str) -> str:
        """Export change log to CSV"""
        df = pd.DataFrame(changes)
        
        # Add metadata
        metadata_row = {
            'change_type': f"# Change Log Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            'product_id': f"Total Changes: {len(changes)}",
            'old_value': '',
            'new_value': '',
            'change_percentage': '',
            'description': '',
            'created_at': ''
        }
        
        df = pd.concat([pd.DataFrame([metadata_row]), df], ignore_index=True)
        
        filepath = os.path.join('exports', f"{filename}.csv")
        df.to_csv(filepath, index=False)
        
        return filepath
    
    def cleanup_old_exports(self, days_old: int = 7):
        """Clean up export files older than specified days"""
        try:
            cutoff_date = datetime.now().timestamp() - (days_old * 24 * 60 * 60)
            exports_dir = 'exports'
            
            if not os.path.exists(exports_dir):
                return
            
            for filename in os.listdir(exports_dir):
                filepath = os.path.join(exports_dir, filename)
                if os.path.isfile(filepath):
                    file_time = os.path.getmtime(filepath)
                    if file_time < cutoff_date:
                        os.remove(filepath)
                        self.logger.info(f"Removed old export file: {filename}")
                        
        except Exception as e:
            self.logger.error(f"Error cleaning up old exports: {e}")
